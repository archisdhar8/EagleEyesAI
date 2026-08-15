from __future__ import annotations

import csv
import io
import os
import time
from datetime import date, datetime, timezone
from typing import Any, Callable
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

import requests
import openpyxl

from . import database


MASSIVE_CONSTITUENTS_URL = "https://api.massive.com/etf-global/v1/constituents"
MASSIVE_TICKERS_URL = "https://api.massive.com/v3/reference/tickers"
ARK_HOLDINGS = {
    "ARKK": "https://assets.ark-funds.com/fund-documents/funds-etf-csv/ARK_INNOVATION_ETF_ARKK_HOLDINGS.csv",
    "ARKF": "https://assets.ark-funds.com/fund-documents/funds-etf-csv/ARK_FINTECH_INNOVATION_ETF_ARKF_HOLDINGS.csv",
    "ARKG": "https://assets.ark-funds.com/fund-documents/funds-etf-csv/ARK_GENOMIC_REVOLUTION_ETF_ARKG_HOLDINGS.csv",
    "ARKW": "https://assets.ark-funds.com/fund-documents/funds-etf-csv/ARK_NEXT_GENERATION_INTERNET_ETF_ARKW_HOLDINGS.csv",
}
INVESCO_HOLDINGS = {
    "QQQ": "https://dng-api.invesco.com/cache/v1/accounts/en_US/shareclasses/QQQ/holdings/fund?idType=ticker&interval=monthly&productType=ETF&loadType=initial",
}
ISHARES_FUNDS = {
    "IVV": ("239726", "ishares-core-sp-500-etf"),
    "IWM": ("239710", "ishares-russell-2000-etf"),
    "AGG": ("239458", "ishares-core-us-aggregate-bond-etf"),
    "EFA": ("239623", "ishares-msci-eafe-etf"),
    "EEM": ("239637", "ishares-msci-emerging-markets-etf"),
    "HYG": ("239565", "ishares-iboxx-high-yield-corporate-bond-etf"),
    "LQD": ("239566", "ishares-iboxx-investment-grade-corporate-bond-etf"),
}
STATE_STREET_FUNDS = {"SPY", "XLC", "XLY", "XLP", "XLE", "XLF", "XLV", "XLI", "XLB", "XLRE", "XLK", "XLU"}
FUND_PROFILES: dict[str, dict[str, Any]] = {
    "QQQ": {"name": "Invesco QQQ Trust", "provider": "Invesco", "source_url": "https://www.invesco.com/qqq-etf/en/about.html", "expense_ratio": .0018},
    "ARKK": {"name": "ARK Innovation ETF", "provider": "ARK Invest", "source_url": "https://www.ark-funds.com/funds/arkk", "expense_ratio": None},
}
POPULAR_FUND_METADATA: dict[str, dict[str, Any]] = {
    "IVV": {"name": "iShares Core S&P 500 ETF", "issuer": "iShares", "asset_class": "Equity", "category": "US large-cap blend", "benchmark": "S&P 500 Index", "expense_ratio": .0003, "source_url": "https://www.ishares.com/us/products/239726/IVV"},
    "VOO": {"name": "Vanguard S&P 500 ETF", "issuer": "Vanguard", "asset_class": "Equity", "category": "US large-cap blend", "benchmark": "S&P 500 Index", "expense_ratio": .0003, "source_url": "https://investor.vanguard.com/investment-products/etfs/profile/voo"},
    "VTI": {"name": "Vanguard Total Stock Market ETF", "issuer": "Vanguard", "asset_class": "Equity", "category": "US total market", "benchmark": "CRSP US Total Market Index", "expense_ratio": .0003, "source_url": "https://investor.vanguard.com/investment-products/etfs/profile/vti"},
    "SPY": {"name": "State Street SPDR S&P 500 ETF Trust", "issuer": "State Street SPDR", "asset_class": "Equity", "category": "US large-cap blend", "benchmark": "S&P 500 Index", "expense_ratio": .000945, "source_url": "https://www.ssga.com/us/en/intermediary/etfs/state-street-spdr-sp-500-etf-trust-spy"},
    "QQQ": {"name": "Invesco QQQ ETF", "issuer": "Invesco", "asset_class": "Equity", "category": "US large-cap growth", "benchmark": "Nasdaq-100 Index", "expense_ratio": .0018, "source_url": "https://www.invesco.com/qqq-etf/en/home.html"},
    "ARKK": {"name": "ARK Innovation ETF", "issuer": "ARK Invest", "asset_class": "Equity", "category": "Disruptive innovation", "strategy": "Active", "benchmark": "Benchmark agnostic", "source_url": "https://www.ark-funds.com/funds/arkk"},
}
ISSUER_PATTERNS = (
    ("iShares", ("ISHARES",)), ("Vanguard", ("VANGUARD",)),
    ("State Street SPDR", ("SPDR",)), ("Invesco", ("INVESCO", "POWERSHARES")),
    ("ARK Invest", ("ARK ",)), ("Charles Schwab", ("SCHWAB",)),
    ("First Trust", ("FIRST TRUST",)), ("VanEck", ("VANECK", "MARKET VECTORS")),
    ("Fidelity", ("FIDELITY",)), ("JPMorgan", ("JPMORGAN", "J.P. MORGAN")),
    ("Dimensional", ("DIMENSIONAL",)), ("Global X", ("GLOBAL X",)),
    ("ProShares", ("PROSHARES",)), ("Direxion", ("DIREXION",)),
)
_LOOKUP_CACHE: dict[str, tuple[float, dict[str, Any]]] = {}
LOOKUP_CACHE_SECONDS = 3600


def _iso_date(value: str) -> str:
    for pattern in ("%Y-%m-%d", "%m/%d/%Y", "%b %d, %Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(value, pattern).date().isoformat()
        except ValueError:
            pass
    return datetime.now(timezone.utc).date().isoformat()


def _safe_provider_error(exc: Exception) -> str:
    if isinstance(exc, requests.HTTPError) and exc.response is not None:
        return f"Provider returned HTTP {exc.response.status_code}"
    return str(exc).split(" for url:", 1)[0][:300]


def _issuer(name: str) -> str | None:
    upper = name.upper()
    for issuer, patterns in ISSUER_PATTERNS:
        if any(pattern in upper for pattern in patterns):
            return issuer
    return None


def _url_with_key(url: str, key: str) -> str:
    parsed = urlparse(url)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query["apiKey"] = key
    return urlunparse(parsed._replace(query=urlencode(query)))


def recognized_fund(ticker: str) -> bool:
    normalized = ticker.upper()
    return normalized in FUND_PROFILES or database.etf_catalog_entry(normalized) is not None


def refresh_etf_catalog() -> dict[str, Any]:
    """Populate the US ETF and single-security ETF reference universe."""
    api_key = os.getenv("POLYGON_API_KEY")
    if not api_key:
        return {"status": "failed", "count": 0, "reason": "Polygon/Massive key is not configured"}
    rows: list[dict[str, Any]] = []
    try:
        for security_type in ("ETF", "ETS"):
            next_url: str | None = MASSIVE_TICKERS_URL
            first = True
            while next_url:
                params = {"market": "stocks", "locale": "us", "active": "true", "type": security_type, "limit": 1000, "sort": "ticker", "apiKey": api_key} if first else None
                target = next_url if first else _url_with_key(next_url, api_key)
                response = requests.get(target, params=params, timeout=20)
                response.raise_for_status()
                payload = response.json()
                observed = datetime.now(timezone.utc).isoformat()
                for item in payload.get("results") or []:
                    ticker, name = str(item.get("ticker") or "").upper(), str(item.get("name") or "").strip()
                    if ticker and name:
                        rows.append({
                            "ticker": ticker, "name": name, "issuer": _issuer(name), "active": bool(item.get("active", True)),
                            "primary_exchange": item.get("primary_exchange"), "currency": str(item.get("currency_name") or "USD").upper(),
                            "provider": "Massive Reference", "source_url": "https://massive.com/docs/rest/stocks/tickers/all-tickers",
                            "effective_at": item.get("last_updated_utc") or observed,
                            "metadata": {"reference_type": item.get("type"), "cik": item.get("cik"), "composite_figi": item.get("composite_figi")},
                        })
                next_url = payload.get("next_url")
                first = False
        count = database.upsert_etf_catalog(rows)
        database.upsert_etf_catalog([
            {"ticker": ticker, **profile, "provider": "Official issuer profile", "effective_at": datetime.now(timezone.utc).isoformat(), "metadata": {"profile_source": "official issuer", "holdings_frequency": "daily" if ticker == "ARKK" else "unknown"}}
            for ticker, profile in POPULAR_FUND_METADATA.items()
        ])
        database.record_etf_refresh("Massive Reference", "catalog", "success", row_count=count, metadata={"security_types": ["ETF", "ETS"]})
        return {"status": "success", "count": count, "provider": "Massive Reference"}
    except (requests.RequestException, ValueError) as exc:
        reason = _safe_provider_error(exc)
        database.record_etf_refresh("Massive Reference", "catalog", "failed", row_count=len(rows), error=reason)
        return {"status": "failed", "count": 0, "reason": reason}


def _massive_snapshot(ticker: str) -> dict[str, Any]:
    api_key = os.getenv("POLYGON_API_KEY")
    if not api_key:
        raise RuntimeError("Polygon/Massive key is not configured")
    response = requests.get(MASSIVE_CONSTITUENTS_URL, params={"composite_ticker": ticker, "limit": 5000, "sort": "effective_date.desc", "apiKey": api_key}, timeout=12)
    if response.status_code == 403:
        raise PermissionError("ETF Global Constituents is not included in the configured market-data plan")
    response.raise_for_status()
    results = response.json().get("results") or []
    if not results:
        raise RuntimeError("The ETF provider returned no constituents")
    latest_date = max(str(item.get("effective_date") or "") for item in results)
    holdings = [{"ticker": str(item.get("constituent_ticker") or "").strip().upper(), "weight": float(item["weight"]), "as_of": latest_date} for item in results if str(item.get("effective_date") or "") == latest_date and item.get("constituent_ticker") and item.get("weight") is not None]
    return {"provider": "Massive ETF Global", "source_url": MASSIVE_CONSTITUENTS_URL, "as_of": latest_date, "holdings": holdings}


def _ark_snapshot(ticker: str) -> dict[str, Any]:
    source_url = ARK_HOLDINGS[ticker]
    response = requests.get(source_url, headers={"User-Agent": "EagleEyes research workspace"}, timeout=12)
    response.raise_for_status()
    holdings, dates = [], []
    for row in csv.DictReader(io.StringIO(response.text.lstrip("\ufeff"))):
        constituent = str(row.get("ticker") or "").strip().upper()
        raw_weight = str(row.get("weight (%)") or "").replace("%", "").strip()
        reported = _iso_date(str(row.get("date") or "").strip())
        if constituent and raw_weight:
            holdings.append({"ticker": constituent, "weight": float(raw_weight) / 100, "as_of": reported, "metadata": {"name": row.get("company")}})
            dates.append(reported)
    if not holdings:
        raise RuntimeError("ARK returned a file without usable holdings")
    return {"provider": "ARK Invest", "source_url": source_url, "as_of": max(dates), "holdings": holdings}


def _invesco_snapshot(ticker: str) -> dict[str, Any]:
    response = requests.get(INVESCO_HOLDINGS[ticker], headers={"Accept": "application/json,*/*", "Referer": "https://www.invesco.com/qqq-etf/en/about.html", "User-Agent": "Mozilla/5.0 EagleEyes"}, timeout=12)
    if response.status_code == 406:
        raise PermissionError("Invesco blocked the server-side holdings request")
    response.raise_for_status()
    payload = response.json()
    as_of = str(payload.get("effectiveDate") or date.today().isoformat())
    holdings = [{"ticker": str(item.get("ticker") or "").strip().upper(), "weight": float(item["percentageOfTotalNetAssets"]) / 100, "as_of": as_of} for item in payload.get("holdings") or [] if item.get("ticker") and item.get("percentageOfTotalNetAssets") is not None]
    if not holdings:
        raise RuntimeError("Invesco returned no usable holdings")
    return {"provider": "Invesco", "source_url": "https://www.invesco.com/qqq-etf/en/about.html", "as_of": as_of, "holdings": holdings}


def _ishares_snapshot(ticker: str) -> dict[str, Any]:
    product_id, slug = ISHARES_FUNDS[ticker]
    source_url = f"https://www.ishares.com/us/products/{product_id}/{slug}/1467271812596.ajax?fileType=csv&fileName={ticker}_holdings&dataType=fund"
    response = requests.get(source_url, headers={"User-Agent": "Mozilla/5.0 EagleEyes"}, timeout=15)
    response.raise_for_status()
    lines = response.text.lstrip("\ufeff").splitlines()
    header_index = next((index for index, line in enumerate(lines) if line.lower().startswith("ticker,")), None)
    if header_index is None:
        raise RuntimeError("iShares returned a file without a holdings header")
    holdings, dates = [], []
    for row in csv.DictReader(lines[header_index:]):
        constituent = str(row.get("Ticker") or "").strip().upper()
        reported = _iso_date(str(row.get("As Of") or row.get("As of") or "").strip())
        try:
            weight = float(str(row.get("Weight (%)") or "").replace(",", "")) / 100
        except ValueError:
            continue
        if constituent and 0 <= weight <= 1:
            holdings.append({"ticker": constituent, "weight": weight, "as_of": reported, "metadata": {"name": row.get("Name")}})
            dates.append(reported)
    if not holdings:
        raise RuntimeError("iShares returned no usable holdings")
    return {"provider": "iShares", "source_url": source_url, "as_of": max(dates), "holdings": holdings}


def _state_street_snapshot(ticker: str) -> dict[str, Any]:
    source_url = f"https://www.ssga.com/us/en/intermediary/library-content/products/fund-data/etfs/us/holdings-daily-us-en-{ticker.lower()}.xlsx"
    response = requests.get(source_url, headers={"User-Agent": "Mozilla/5.0 EagleEyes"}, timeout=15)
    response.raise_for_status()
    workbook = openpyxl.load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    date_cell = next((str(row[1]) for row in rows[:8] if row and row[0] == "Holdings:" and len(row) > 1), "")
    as_of = _iso_date(date_cell.replace("As of ", "").strip())
    header_index = next((index for index, row in enumerate(rows) if row and row[0] == "Name" and len(row) > 4 and row[1] == "Ticker"), None)
    if header_index is None:
        raise RuntimeError("State Street returned a workbook without a holdings header")
    holdings = []
    for row in rows[header_index + 1:]:
        if not row or len(row) < 5:
            continue
        constituent = str(row[1] or "").strip().upper()
        try:
            weight = float(row[4]) / 100
        except (TypeError, ValueError):
            continue
        if constituent and 0 <= weight <= 1:
            holdings.append({"ticker": constituent, "weight": weight, "as_of": as_of, "metadata": {"name": row[0], "reported_sector": row[5] if len(row) > 5 else None}})
    if not holdings:
        raise RuntimeError("State Street returned no usable holdings")
    return {"provider": "State Street SPDR", "source_url": source_url, "as_of": as_of, "holdings": holdings}


def _profile(ticker: str) -> dict[str, Any]:
    catalog = database.etf_catalog_entry(ticker) or {}
    return FUND_PROFILES.get(ticker, {"name": catalog.get("name") or ticker, "provider": catalog.get("issuer") or catalog.get("provider") or "ETF issuer", "source_url": catalog.get("source_url"), "expense_ratio": catalog.get("expense_ratio")})


def holdings_freshness(as_of: str | None, expected_frequency: str | None = None) -> dict[str, Any]:
    if not as_of:
        return {"status": "unavailable", "label": "Holdings unavailable", "days_old": None}
    try:
        age = max(0, (date.today() - datetime.fromisoformat(as_of[:10]).date()).days)
    except ValueError:
        return {"status": "unavailable", "label": "Holdings date unavailable", "days_old": None}
    if expected_frequency == "daily" and age <= 2:
        return {"status": "daily", "label": "Daily holdings", "days_old": age}
    if age <= 45:
        return {"status": "delayed", "label": "Delayed holdings", "days_old": age}
    return {"status": "stale", "label": "Stale holdings", "days_old": age}


def ensure_fund_data(ticker: str, force: bool = False) -> dict[str, Any]:
    normalized = ticker.strip().upper()
    if not recognized_fund(normalized):
        return {"status": "not_applicable", "reason": "The symbol is not in the US ETF catalog."}
    existing = database.fund_reference_data([normalized])
    if not force and existing.get("funds") and existing.get("holdings"):
        latest = existing["holdings"][0].get("as_of")
        return {"status": "available", "provider": existing["funds"][0].get("provider"), "reason": "Using the latest stored ETF snapshot.", "freshness": holdings_freshness(latest, (database.etf_catalog_entry(normalized) or {}).get("metadata", {}).get("holdings_frequency"))}
    cached = _LOOKUP_CACHE.get(normalized)
    if not force and cached and time.monotonic() - cached[0] < LOOKUP_CACHE_SECONDS:
        return cached[1]
    profile, errors, snapshot = _profile(normalized), [], None
    adapters: list[Callable[[str], dict[str, Any]]] = [_massive_snapshot]
    if normalized in ARK_HOLDINGS: adapters.append(_ark_snapshot)
    if normalized in INVESCO_HOLDINGS: adapters.append(_invesco_snapshot)
    if normalized in ISHARES_FUNDS: adapters.append(_ishares_snapshot)
    if normalized in STATE_STREET_FUNDS: adapters.append(_state_street_snapshot)
    for adapter in adapters:
        try:
            snapshot = adapter(normalized)
            break
        except (requests.RequestException, RuntimeError, PermissionError, ValueError) as exc:
            errors.append(_safe_provider_error(exc))
    if snapshot is None:
        result = {"status": "missing", "provider": profile["provider"], "reason": "; ".join(dict.fromkeys(errors)) or "The issuer has no connected machine-readable holdings feed.", "source_url": profile.get("source_url"), "freshness": holdings_freshness(None)}
        database.record_etf_refresh(str(profile["provider"]), "holdings", "failed", normalized, error=result["reason"])
        _LOOKUP_CACHE[normalized] = (time.monotonic(), result)
        return result
    database.upsert_etf_catalog([{"ticker": normalized, "name": profile["name"], "issuer": profile["provider"], "expense_ratio": profile.get("expense_ratio"), "provider": snapshot["provider"], "source_url": snapshot["source_url"], "effective_at": snapshot["as_of"], "holdings_count": len(snapshot["holdings"]), "metadata": {"holdings_frequency": "daily" if normalized in ARK_HOLDINGS or normalized in STATE_STREET_FUNDS else "unknown"}}])
    database.save_fund_reference_snapshot(normalized, profile.get("expense_ratio"), snapshot["provider"], snapshot["source_url"], snapshot["as_of"], snapshot["holdings"], metadata={"name": profile["name"], "retrieval": "on-demand ETF research"})
    database.rebuild_etf_sector_exposures(normalized, snapshot["as_of"], snapshot["source_url"])
    database.record_etf_refresh(snapshot["provider"], "holdings", "success", normalized, len(snapshot["holdings"]), metadata={"as_of": snapshot["as_of"]})
    result = {"status": "available", "provider": snapshot["provider"], "reason": f"Loaded {len(snapshot['holdings'])} holdings as of {snapshot['as_of']}.", "source_url": snapshot["source_url"], "freshness": holdings_freshness(snapshot["as_of"], "daily" if normalized in ARK_HOLDINGS or normalized in STATE_STREET_FUNDS else None)}
    _LOOKUP_CACHE[normalized] = (time.monotonic(), result)
    return result
